from fastapi import UploadFile
import openpyxl
from io import BytesIO
import os
from pathlib import Path
from functions.functions_database import ObtenerActividaEconomica

async def InfoInsercionTable(file: UploadFile):

    workbook = openpyxl.load_workbook(BytesIO(await file.read()))

    hoja = workbook.active

    filas = []

    primeraF = True

    for fila in hoja.iter_rows():
        valores_fila = []
        if primeraF:
            primeraF = False
            continue

        for celda in fila:
            valor = celda.value

            if valor == '':
                valor = None

            valores_fila.append(valor)

        filas.append(valores_fila)

    return filas


def GenerarExcelIcsoPromotor(Resultado,CedulaPromo):
    libro = openpyxl.Workbook()
    hoja = libro.active

    encabezado = [
        "nombres", "primerapellido", "segundoapellido", "tipoidentificacion", "identificacion", "montoaprobado", "cod",
        "promotor", "zona", "linea", "convenio", "fechanacimiento", "sexo", "numerocuentabancaria", "entidad", "direccion", "barrio", "estrato",
        "telefono", "celular", "correo", "personascargo", "tipodevivienda", "nivelescolar", "estadocivil", "nombresdelconyuge", "primerapellidodelconyuge",
        "segundoapellidodelconyuge", "tipodocumentoconyuge", "numerodocumentoconyuge", "telefonoconyuge", "lgbtiq", "discapacitado", "victima", "poblacioncampesina",
        "cabezahogar", "etnia", "estadolaboral", "nitempresa", "nombreempresa", "telefonoempresa", "tiemponegocio", "periodo", "direccionnegocio", "barrionegocio",
        "sectoreconomico", "actividadeconomica", "nombrereferencia1", "telefonoreferencia1", "nombrereferencia2", "telefonoreferencia2", "nombrefdm1", "primerapellidofdm1",
        "segundoapellidofdm1", "parentescofdm1", "tipodocumentofdm1", "documentofdm1", "telefonofdm1", "participacionFDM1", "nombrefdm2", "primerapellidofdm2", "segundoapellidofdm2",
        "parentescofdm2", "tipodocumentofdm2", "documentofdm2", "telefonofdm2", "participacionfdm2"
    ]

    hoja.append(encabezado)
    Promotor = 0
    for fila in Resultado:
        if fila[5] == None:
            nombres = f"{fila[4]} " #
        else:
            nombres = f"{fila[4]} {fila[5]}" #

        primerA = fila[2] #
        segundoA = fila[3] #
        texto1 = "CEDULA"
        identificacion = fila[1]
        montoAprobado = fila[20]
        codigo = fila[34]
        cedula_promotor = fila[32]
        zona = fila[35]
        linea = str(fila[25])[0]
        convenio = 99
        fecha_nacimiento = fila[10]
        sexo = fila[48]
        if sexo == "F":
            sexo = "FEMENI"
        else:
            sexo = "MASCUL"

        cuenta_bancaria = fila[37]
        entidad = "002"
        direccion = fila[7]
        barrio = fila[50]
        estrato = fila[15]
        telefono = fila[8]
        if telefono == None:
            telefono = ""
        else:
            telefono = f"604{telefono}"
        
        celular = fila[16]
        correo = fila[40]
        personascargo = 99
        tipovivienda = ""
        nivelescolar = "PRIMAR"
        estadocivil = "SOLTER"
        nombres_conyugue = ""
        primerA_conyugue = ""
        segundoA_conyugue = ""
        tipo_documento_conyugue = "CEDULA"
        numero_documento_conyugue = 9999999999
        telefono_conyugue = 9999999999
        lgbtiq = "No"
        discapacitado = "No"
        victima = "No"
        poblacioncampesina = "No"
        cabezahogar = "No"
        etnia = "Otro"
        estadolaboral = "EMPREN"
        nitempresa = ""
        nombreempresa = ""
        telefonoempresa = ""
        tiemponegocio = 6
        periodo = "MESES"
        direccionnegocio = direccion
        barrionegocio = barrio

        Economico = ObtenerActividaEconomica(fila[18])
        sectoreconomico = Economico[1]
        actividadeconimica = Economico[0]
        nombrereferencia1 = "SIN NOMBRE"
        telefonoreferencia1 = 9999999999
        nombrereferencia2 = "SIN NOMBRE"
        telefonoreferencia2 = 9999999999

        nombrefdm1 = fila[42]
        if nombrefdm1 == None:
            nombrefdm1 = "SIN NOMBRE"
        primerapellidofdm1 = nombrefdm1
        segundoapellidofdm1 = ""
        parentescofdm1 = fila[43]
        tipodocumentofdm1 = "CEDULA"
        documentofdm1 = fila[41]
        telefonofdm1 = 9999999999
        participacionfdm1 = 100

        nombrefdm2 = ""
        primerapellidofdm2 = ""
        segundoapellidofdm2 = ""
        parentescofdm2 = ""
        tipodocumentofdm2 = ""
        documentofdm2 = ""
        telefonofdm2 = ""
        participacionfdm2 = ""
  
        hoja.append([
            nombres, primerA, segundoA, texto1, identificacion, montoAprobado, codigo, cedula_promotor,
            zona, linea, convenio, fecha_nacimiento, sexo, cuenta_bancaria, entidad, direccion, barrio, estrato,
            telefono, celular, correo, personascargo, tipovivienda, nivelescolar, estadocivil, nombres_conyugue,
            primerA_conyugue, segundoA_conyugue, tipo_documento_conyugue, numero_documento_conyugue, telefono_conyugue,
            lgbtiq, discapacitado, victima, poblacioncampesina, cabezahogar, etnia, estadolaboral, nitempresa,
            nombreempresa, telefonoempresa, tiemponegocio, periodo, direccionnegocio, barrionegocio, sectoreconomico,
            actividadeconimica, nombrereferencia1, telefonoreferencia1, nombrereferencia2, telefonoreferencia2, nombrefdm1,
            primerapellidofdm1, segundoapellidofdm1, parentescofdm1, tipodocumentofdm1, documentofdm1, telefonofdm1,
            participacionfdm1, nombrefdm2, primerapellidofdm2, segundoapellidofdm2, parentescofdm2, tipodocumentofdm2,
            documentofdm2, telefonofdm2, participacionfdm2
        ])
        

    descargas_carpeta = str(Path.home() / "Descargas")
    
    os.makedirs(descargas_carpeta,exist_ok=True)

    archivo = os.path.join(descargas_carpeta,"Archivo_icso_promotor_{}.xlsx".format(CedulaPromo))
    libro.save(archivo)

    return archivo
