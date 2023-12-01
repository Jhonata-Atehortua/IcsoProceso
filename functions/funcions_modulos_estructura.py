from fastapi import UploadFile
import openpyxl
from io import BytesIO
import os
from datetime import datetime
from pathlib import Path
from functions.functions_database import EstructuraModuloCedula

async def ObtenerCedulaBuscar(file: UploadFile):

    workbook = openpyxl.load_workbook(BytesIO(await file.read()))

    hoja = workbook.active

    filas = []

    primeraF = True

    for fila in hoja.iter_rows():
        valores_fila = []
        if primeraF:
            primeraF = False
            continue

        for celda in fila:
            valor = celda.value

            if valor == '':
                valor = None

            valores_fila.append(valor)

        filas.append(valores_fila)

    return filas


def GenerarEstructuraModulo(Cedulas):
    libro = openpyxl.Workbook()
    hoja = libro.active

    encabezado = [
        "NIT", "PRIMER APELLIDO", "SEGUNDO APELLIDO", "NOMBRES", "SEGUNDO NOMBRE", "NOMBRE INTEGRADO", "DIRECCION",
        "FECHA NACIMIENTO", "EDAD", "CODIGO COMUNA", "ESTRATO", "CELULAR", "CIRCULO" ,"NOMBRE CIRCULO","MUNICIPIO"
    ]

    hoja.append(encabezado)

    for registro in Cedulas:
        resultado = EstructuraModuloCedula(registro[0])
        if len(resultado) > 0:
            cedula = resultado[0][1]
            primerA = resultado[0][2]
            segundoA = resultado[0][3]
            primerN = resultado[0][4]
            segunN = resultado[0][5]
            nombreI = resultado[0][6]
            direccion = resultado[0][7]
            fechaN = resultado[0][10]
            fechaN = resultado[0][10]
            fechaAc = datetime.now().date()
            Edad =  fechaAc.year -fechaN.year
            codigoC = resultado[0][13]
            estrato = resultado[0][15]
            celular = resultado[0][16]
            circulo = resultado[0][25]
            nombreC = resultado[0][31]
            municipio = resultado[0][36]
            hoja.append([
                cedula,primerA,segundoA,primerN,segunN,nombreI,direccion,fechaN,Edad,codigoC,estrato,celular,circulo,nombreC,municipio
            ])
            continue    
        else:
            hoja.append([registro[0]])
   
    descargas_carpeta = str(Path.home() / "Descargas")
    
    os.makedirs(descargas_carpeta,exist_ok=True)

    archivo = os.path.join(descargas_carpeta,"Modulo.xlsx")
    libro.save(archivo)

    return archivo
